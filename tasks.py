import os
import shutil
import pandas
import numpy as np
import re
import openpyxl
from datetime import datetime, timedelta
from helper import WriteToExcel, CleanServiceName, Employee, Commission
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from weasyprint import HTML
import re
import openpyxl
from openpyxl.styles import Alignment, Font
from jinja2 import Environment, FileSystemLoader
from database.salary_slip_manager import SalarySlipManager
from decimal import Decimal

# Read file excel and collect data with sheetnama rptOmzet, get col from A-H except D and skip 2 frist row
object_data = pandas.read_excel('Excel/omzet.xlsx', sheet_name='rptOmzet', usecols='A:B, E:H', skiprows=2)
# Get index data with non NaN value
index_names = object_data.Employee.dropna().index
# Initiate last row data
last_row = object_data.Employee.index[-1]

writer = pandas.ExcelWriter('/Users/rusadi/Projects/extra/excel-pandas/Excel/Final Omzet.xlsx', engine='xlsxwriter')

# Loop index names to get data omzet per employee
def Calculation(position, omzet, reduction, commission, idx_item, total, price, disc_value):
    if position == 'Stylist':
        if int(disc_value) > 10:
            calculateCommission(omzet, reduction, commission, idx_item, price) 
        else:
            calculateCommission(omzet, reduction, commission, idx_item, total) 
    else:
        if int(disc_value) > 10:
            calCommission(omzet, commission, idx_item, price) 
        else:
            calCommission(omzet, commission, idx_item, total)

def calCommission(omzet, commission, idx_item, nominal):
    omzet.loc[idx_item, 'Nett'] = (nominal * commission) / 100
    omzet.loc[idx_item, '%'] = (omzet.loc[idx_item, 'Nett'] / nominal) * 100

def calculateCommission(omzet, reduction, commission, idx_item, nominal):
    potongan = omzet.loc[idx_item, 'Potongan'] = (nominal * reduction) / 100
    bruto = omzet.loc[idx_item, 'Bruto'] = nominal - potongan
    nett = omzet.loc[idx_item, 'Nett'] = (bruto * commission) / 100
    omzet.loc[idx_item, 'potongan %'] = (potongan / nominal) * 100 
    omzet.loc[idx_item, 'komisi %'] = (nett / bruto) * 100

def generateCommission(position, omzet, index_product, last_row_product, idx, last, target, service_name, category, reduction, commission):
    if target == last:
        omzet.loc[target:last_row_product, 'Description'] = service_name
        omzet.loc[target:last_row_product, 'Category'] = category
        array_omzet = omzet.loc[target:last_row_product].index
        CommissioerItem(position, omzet, reduction, commission, array_omzet)
    else:
        next = index_product[idx + 1] - 1
        omzet.loc[target:next, 'Description'] = service_name
        omzet.loc[target:next, 'Category'] = category
        array_omzet = omzet.loc[target:next].index
        CommissioerItem(position, omzet, reduction, commission, array_omzet)

def CommissioerItem(position, omzet, reduction, commission, array_omzet):
    for idx_item in array_omzet:
        total = omzet.loc[idx_item, 'Total Nett']
        price = omzet.loc[idx_item,'Price']
        discount = omzet.loc[idx_item,'Total Disc.']
        disc_value = (discount / price) * 100
                
        Calculation(position, omzet, reduction, commission, idx_item, total, price, disc_value)

def getDataOmzet(object_data, index_names, last_row, index, last_idx, target_idx):
    if target_idx == last_idx:
        # get data omzet each employee in last index of employee from range row and range column
        omzet_employees = object_data.iloc[target_idx:last_row + 1, 0:7]
    else:
        # get data omzet each employe form range row and range column
        next_index = index_names[index + 1]
        omzet_employees = object_data.iloc[target_idx:next_index, 0:7]
    return omzet_employees

def regenerateDataOmzet(object_data, index_names, last_row):
    for index in range(len(index_names)):
        last_idx = index_names[-1]
        target_idx = index_names[index]

        omzet_employees = getDataOmzet(object_data, index_names, last_row, index, last_idx, target_idx)
    
        # Initiate employee name convert to title case 
        sheet_name = (omzet_employees['Employee'].values[0]).title()
        position = Employee()[sheet_name.lower()].title()
    
        print(sheet_name)
        print(position)
        print('-------------')
    
        omzet = omzet_employees.iloc[0:omzet_employees.shape[0], 1:omzet_employees.shape[1]]
        # get index service name except NaN value from column Descriptions
        index_product = omzet.Description.dropna().index
        # initiate last index from loop service name
        last_row_product = omzet.Description.index[-1]
        
        # Loop index service name
        for idx in range(len(index_product)):
            last = index_product[-1]
            target = index_product[idx]
            
            # Clean name of services
            service_name = CleanServiceName(omzet.Description[target])['service_name']
            category = CleanServiceName(omzet.Description[target])['category'].lower()
            reduction = Commission()[position.lower()][category]['reduction']
            commission = Commission()[position.lower()][category]['commission']
            
            if sheet_name == 'Saini' and category == 'mp':
                commission = int(50)
                
            if sheet_name == 'Marni' and category == 'mp':
                commission = int(25)
                
            if sheet_name == 'Febriana' and category == 'mp':
                commission = int(25)
                
            if sheet_name == 'Tatang' and (category == 'blow' or category == 'haircut'):
                reduction = int(20)
            
            if sheet_name == 'Tatang' and category == 'chemical':
                reduction = int(40)
            
            if sheet_name == 'Nia ' and category == 'makeup':
                commission = int(40)
        
            # Rename NaN value with service name
            generateCommission(position, omzet, index_product, last_row_product, idx, last, target, service_name, category, reduction, commission)
            
        #  Remove parent service name
        omzet.drop(index_product, axis=0, inplace=True)
        # Sort data ascending by column description
        omzet.sort_values(by=['Description'], ascending=True, inplace=True)
        # write to excel with each sheetname by employee name
        destination =  "/Users/rusadi/Projects/extra/excel-pandas/Excel/omzet.xlsx"
        WriteToExcel(destination, omzet, sheet_name, False)
        
def generateFinalOmzet():
    df_omzet = pandas.read_excel('Excel/omzet.xlsx', sheet_name=None)
    actual_sheets = [sheet for sheet in df_omzet if sheet != 'rptOmzet']
    data_omzet = pandas.read_excel('Excel/Payroll.xlsx', sheet_name='Omzet', usecols=['Nama','Omzet','Komisi', 'Bonus Omzet'])
    data_omzet.rename(index=data_omzet.Nama, inplace=True)
    total_omzet = 0

    for item in actual_sheets:
        data_sheet = pandas.read_excel('Excel/omzet.xlsx', sheet_name=item)

        grouped = data_sheet.groupby('Description').sum()
        grouped.loc['Grand Total'] = grouped.sum()
        grouped.insert(0,'Description', grouped.index)
        
        if 'potongan %' in grouped.columns.values:
            grouped.drop(columns=['potongan %', 'komisi %', 'Category'], inplace=True)
            grouped.rename(columns = {'Description':'SERVICE', 'Price':'PRICE', 'Total Qty':'QTY', 'Total Disc.':'DISC', 'Total Nett':'AFTER DISC', 'Potongan':'POTONGAN', 'Bruto':'BRUTO', 'Nett':'NETT'}, inplace = True)
        else:
            grouped.drop(columns=['%', 'Category'], inplace=True)
            grouped.rename(columns = {'Description':'SERVICE', 'Price':'PRICE', 'Total Qty':'QTY', 'Total Disc.':'DISC', 'Total Nett':'BRUTO', 'Nett':'NETT'}, inplace = True)
            
        total_omzet = grouped.values[-1][4]
        total_komisi = grouped.values[-1][-1]
        data_omzet.loc[item.strip(), 'Omzet'] = total_omzet
        data_omzet.loc[item.strip(), 'Komisi'] = total_komisi
        
        # Create a Pandas Excel writer using XlsxWriter engine.
        grouped.to_excel(writer, sheet_name=item, startrow=4, index=False)
        last_index =  grouped.shape[0] + 4
        
        # Get workbook and worksheet objects
        workbook  = writer.book
        worksheet = writer.sheets[item]
        max_col = grouped.shape[1]
        
        # Format cell
        merge_format = workbook.add_format({'bold': True, 'align':'center'})
        center_format = workbook.add_format({'align':'center'})
        format_number = workbook.add_format({'num_format': '#,##0'})

        # Set merge_range by length of colums names
        len_cols = len(grouped.columns)
        current_month = datetime.now().strftime("%B")
        previous_month = (datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)).strftime("%B")
        year = datetime.now().year
        
        # Merge row title
        worksheet.merge_range(0, 0, 0, len_cols - 1, 'T-Style Salon'.upper(), merge_format)
        worksheet.merge_range(1, 0, 1, len_cols - 1, f'Periode 21 { previous_month } - 20 { current_month } { year }'.upper(), merge_format)
        worksheet.merge_range(2, 0, 2, len_cols - 1, item.upper(), merge_format)
        
        # Format column
        worksheet.set_row(4, None, center_format)
        worksheet.set_column('A:A',30)
        worksheet.set_column(f'B6:B{last_index}',10,format_number)
        worksheet.set_column(f'D6:H{last_index}',10, format_number)
        worksheet.set_column(f'C6:C{last_index}',8, center_format)
        
        # Set style on table
        col_settings = [{ 'header': column } for column in grouped.columns]    
        worksheet.add_table(4, 0, last_index, max_col - 1, { 'autofilter': False, 'columns': col_settings, 'style': 'Table Style Light 11', 'total_row': True })

    writer.close()

# Helper function to create the output folder dynamically for either "Slip Gaji" or "Omzet"
def create_output_folder(folder_type='Slip Gaji'):
    output_folder = f'PDF/{datetime.today().strftime("%Y-%m-%d")}/{folder_type}'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
    os.makedirs(output_folder)
    return output_folder

# Helper function to load Excel workbook
def load_workbook_from_excel(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File {file_path} tidak ditemukan.")
    if os.path.getsize(file_path) == 0:
        raise ValueError(f"File {file_path} kosong atau corrupt.")
    return openpyxl.load_workbook(file_path)

# Helper function to format numeric columns in DataFrame
def format_numeric_columns(df, numeric_columns):
    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: f"{x:,.0f}" if pandas.notnull(x) else '')
    return df

# Helper function to sanitize sheet name
def sanitize_sheet_name(sheet_name):
    return re.sub(r'[\\/*?:"<>|]', "", sheet_name.strip()) + ' Omzet'

# Helper function to generate HTML content from DataFrame and sheet data
def generate_html_content(df, sheet_name, row_1_text, row_2_text, row_3_text, column_widths):
    html_content = f'''
    <html>
    <head>
    <style>
        @page {{
            size: A4 portrait;
            margin: 0cm;
        }}
        body {{
            font-family: Arial, sans-serif;
            margin: 1cm;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 10px;
        }}
        .header-title {{
            text-align: center;
            font-weight: bold;
            font-size: 14px;
            margin-bottom: 5px;
        }}
        th, td {{
            padding: 4px;
            text-align: left;
            border: 1px solid #e0e0e0;
            word-wrap: break-word;
        }}
        th {{
            background-color: #D9EAD3;
            font-weight: bold;
            color: #333;
            text-align: center;
        }}
        tr:nth-child(odd) {{
            background-color: #FFFFFF;
        }}
        tr:nth-child(even) {{
            background-color: #F9F9F9;
        }}
        .sheet-title {{
            font-size: 24px;
            margin-top: 20px;
            margin-bottom: 10px;
        }}
        .center {{
            text-align: center;
        }}
        .bold {{
            font-weight: bold;
        }}
        .format-number {{
            text-align: right;
        }}
        .qty {{
            text-align: center;
            width: {column_widths.get('QTY', 10)}ch;
        }}
        .service {{
            text-align: left;
            width: {column_widths.get('SERVICE', 10)}ch;
        }}
        .price, .disc, .after-disc, .bruto, .nett {{
            text-align: right;
            width: {column_widths.get('PRICE', 10)}ch;
        }}
    </style>
    </head>
    <body>
        <div class="header-title">{row_1_text}</div>
        <div class="header-title">{row_2_text}</div>
        <div class="header-title">{row_3_text}</div>
        <br></br>
        <table class="table table-bordered">
    '''

    # Add table headers
    html_content += '<thead><tr>'
    for col in df.columns:
        html_content += f'<th>{col}</th>'
    html_content += '</tr></thead><tbody>'

    # Add table rows
    for index, row in df.iterrows():
        row_class = 'class="bold"' if 'Grand Total' in row.values else ''
        html_content += f'<tr {row_class}>'
        for col in df.columns:
            align_class = 'class="qty"' if col == 'QTY' else 'class="service"' if col == 'SERVICE' else 'class="format-number"'
            html_content += f'<td {align_class}>{row[col]}</td>'
        html_content += '</tr>'

    # Close table and body
    html_content += '</tbody></table></body></html>'
    return html_content

# Main function to generate PDF
def generatePDF():
    base_path = os.path.dirname(os.path.abspath(__file__))
    excel_filename = 'Excel/Final Omzet.xlsx'
    excel_file = os.path.join(base_path, excel_filename)
    
    # Load Excel workbook
    workbook = load_workbook_from_excel(excel_file)
    
    # Create output folder
    output_folder = create_output_folder('Omzet')

    # Loop through each sheet and generate PDF
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Convert worksheet to DataFrame
        data = sheet.iter_rows(min_row=6, values_only=True)
        columns = [cell.value for cell in sheet[5]]
        df = pandas.DataFrame(data, columns=columns)

        # Format numeric columns
        numeric_columns = ['PRICE', 'QTY', 'DISC', 'AFTER DISC', 'NETT', 'POTONGAN']
        df = format_numeric_columns(df, numeric_columns)

        # Sanitize sheet name for filename
        safe_sheet_name = sanitize_sheet_name(sheet_name)

        # Get row data for header
        row_1 = [cell.value for cell in sheet[1]]
        row_2 = [cell.value for cell in sheet[2]]
        row_3 = [cell.value for cell in sheet[3]]
        row_1_text = next((text for text in row_1 if text), "")
        row_2_text = next((text for text in row_2 if text), "")
        row_3_text = next((text for text in row_3 if text), "")

        # Calculate column widths
        column_widths = {col: max(df[col].apply(lambda x: len(str(x)) if pandas.notnull(x) else 0).max(), len(col)) + 2 for col in df.columns}

        # Generate HTML content
        html_content = generate_html_content(df, sheet_name, row_1_text, row_2_text, row_3_text, column_widths)

        # Output PDF path
        output_pdf_path = os.path.join(output_folder, f'{safe_sheet_name}.pdf')

        # Convert HTML to PDF using WeasyPrint
        HTML(string=html_content).write_pdf(output_pdf_path, presentational_hints=True)
        
        print(f'✅ PDF berhasil dibuat: {output_pdf_path}')

# Helper function to format decimal values to string with thousands separator
def format_decimal(value, default='0'):
    return f"{value:,.0f}" if isinstance(value, Decimal) else default

# Helper function to fetch salary items
def get_salary_items(manager, slip_id):
    manager.cursor.execute("SELECT * FROM salary_item WHERE slip_id = %s", (slip_id,))
    items = manager.cursor.fetchall()
    income_items = {item['item_name']: item['amount'] for item in items if item['item_type'] == 'income'}
    deduction_items = {item['item_name']: item['amount'] for item in items if item['item_type'] == 'deduction'}
    return income_items, deduction_items

# Helper function to fetch attendance data
def get_attendance(manager, slip_id):
    manager.cursor.execute("SELECT * FROM attendance_record WHERE slip_id = %s", (slip_id,))
    return manager.cursor.fetchone()

# Helper function to render salary slip as HTML
def render_salary_slip_html(staff, income_items, deduction_items, slip, attendance, template):
    html_data = {
        'tanggal': datetime.today().strftime('%B %Y'),
        'nama': staff['full_name'].upper(),
        'jabatan': staff['position'].upper(),
        'gaji': format_decimal(income_items.get('Gaji Pokok', Decimal('0'))),
        'komisi': format_decimal(income_items.get('Komisi', Decimal('0'))),
        'lembur': format_decimal(income_items.get('Lembur', Decimal('0'))),
        'bpjs': format_decimal(income_items.get('BPJS', Decimal('0'))),
        'prestasi_absensi': format_decimal(income_items.get('Prestasi Absensi', '-')),
        'prestasi_omzet': format_decimal(income_items.get('Prestasi Omzet', '-')),
        'total_pendapatan': format_decimal(slip['total_income']),
        'telat': format_decimal(deduction_items.get('Telat', Decimal('0'))),
        'kasbon': format_decimal(deduction_items.get('Kasbon', '-')),
        'sisa_kasbon': format_decimal(deduction_items.get('Sisa Kasbon', '')),
        'total_potongan': format_decimal(slip['total_deduction']),
        'total_salary': format_decimal(slip['net_salary']),
        'total_hadir': attendance.get('hadir', 0) if attendance else 0,
        'total_sakit': attendance.get('sakit', 0) if attendance else 0,
        'total_izin': attendance.get('izin', 0) if attendance else 0,
        'total_cuti': attendance.get('cuti', 0) if attendance else 0,
        'total_lembur': attendance.get('lembur', 0) if attendance else 0,
        'total_telat': attendance.get('telat', 0) if attendance else 0,
    }
    return template.render(html_data)

# Helper function to generate the PDF from rendered HTML
def generate_pdf_from_html(rendered_html, output_path):
    base_url = os.path.join(os.getcwd(), 'images')
    HTML(string=rendered_html, base_url=base_url).write_pdf(output_path)
    print(f"✅ Slip gaji dibuat: {output_path}")

# Main function to generate salary slips
def generate_salary_slips():
    manager = SalarySlipManager()
    try:
        output_folder = create_output_folder('Slip Gaji')

        # Load the template
        env = Environment(loader=FileSystemLoader('template'))
        template = env.get_template('slip_gaji.html')

        # Fetch all active staff
        manager.cursor.execute("SELECT * FROM staff WHERE status = 'Active'")
        staffs = manager.cursor.fetchall()

        for staff in staffs:
            staff_id = staff['id']

            # Fetch the latest salary slip for the staff
            manager.cursor.execute("SELECT * FROM salary_slip WHERE staff_id = %s ORDER BY id DESC LIMIT 1", (staff_id,))
            slip = manager.cursor.fetchone()
            if not slip:
                continue

            slip_id = slip['id']

            # Fetch salary items (income and deduction)
            income_items, deduction_items = get_salary_items(manager, slip_id)

            # Fetch attendance record
            attendance = get_attendance(manager, slip_id)

            # Render the HTML content for the salary slip
            rendered_html = render_salary_slip_html(staff, income_items, deduction_items, slip, attendance, template)

            # Output file name and path
            filename = f"{staff['full_name'].replace(' ', '_')}.pdf"
            output_path = os.path.join(output_folder, filename)

            # Generate PDF from the rendered HTML
            generate_pdf_from_html(rendered_html, output_path)

    finally:
        manager.close()

# regenerateDataOmzet(object_data, index_names, last_row)
generateFinalOmzet()
generatePDF()
generate_salary_slips()